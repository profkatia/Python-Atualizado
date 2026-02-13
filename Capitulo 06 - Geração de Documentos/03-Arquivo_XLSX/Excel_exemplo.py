#Importação da biblioteca openpyxl
# from openpyxl import load_workbook

# # # Caminho do arquivo Excel
# caminho_arquivo=r"03-Arquivo_XLSX\Base_Dados Vendedor.xlsx"

# # # Carrega o arquivo
# workbook=load_workbook(filename=caminho_arquivo)

# # Seleciona a planilha ativa (ou use workbook['NomeDaPlanilha'] para uma específica)
# planilha=workbook.active
# # # Itera pelas linhas e colunas
# for linha in planilha.iter_rows(values_only=True):
#     print(linha)

#---------------------------------------------------------------------------------------
from openpyxl import load_workbook

# Caminho do arquivo
caminho_arquivo = r"03-Arquivo_XLSX\Base_Dados Vendedor.xlsx"

# Abre o arquivo
workbook = load_workbook(filename=caminho_arquivo)

# Seleciona a planilha ativa
planilha = workbook.active

# Dados a serem adicionados (nova linha)
nova_linha = ["Carlos Silva", "SP", 8500, 0.10]

# Adiciona a linha no final da planilha
planilha.append(nova_linha)

# Salva o arquivo
workbook.save(caminho_arquivo)

print("Nova linha adicionada com sucesso!")


